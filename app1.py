import streamlit as st
import pandas as pd
import openpyxl
import re
import io
import time
import os
import streamlit.components.v1 as components

# --- 1. CONFIGURATION DE LA PAGE ---
st.set_page_config(page_title="Harness Data Engine", layout="wide", page_icon="🚀")

if "df_principal" not in st.session_state:
    st.session_state.df_principal = None

# Initialisation de la mémoire session
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "welcome_done" not in st.session_state:
    st.session_state.welcome_done = False

# --- 2. FONCTION DE SCROLL (Moteur JS) ---
def auto_scroll():
    """Force le scroll vers l'ancre invisible en bas de page"""
    components.html(
        """
        <script>
            setTimeout(function() {
                var anchor = window.parent.document.getElementById("end-of-chat");
                if (anchor) {
                    anchor.scrollIntoView({ behavior: "smooth" });
                }
            }, 150); // Délai pour laisser le texte s'afficher
        </script>
        """,
        height=0,
    )

# --- 3. LE MOTEUR DE DONNÉES ---
class StreamlitDataEngine:
    def __init__(self):
        self.variables = {}
        self.df = None
        self.file_obj = None

    def add_message(self, text):
        """Affiche le message lettre par lettre et déclenche le scroll"""
        with st.chat_message("assistant"):
            placeholder = st.empty()
            full_response = ""
            for char in text:
                full_response += char
                placeholder.markdown(full_response + "▌")
                time.sleep(0.008) # Vitesse rapide et fluide
            placeholder.markdown(full_response)
        
        # Sauvegarde et Scroll
        st.session_state.chat_history.append(text)
        auto_scroll()

    def get_param(self, param_str, key):
        try:
            parts = [p.strip() for p in str(param_str).split(',')]
            for p in parts:
                if p.startswith(key):
                    return p.split(':', 1)[1].strip().replace('"', '').replace('[', '').replace(']', '')
            return None
        except: return None
      
    def match_col(self, pattern, col_list):
        p_clean = pattern.replace('*', '').lower().strip()
        for col in col_list:
            col_str = str(col).lower().strip()
            if pattern.endswith('*') and col_str.startswith(p_clean): return col
            elif col_str == p_clean: return col
        return None

    def run_pipeline(self, rules_df, uploaded_file):
        self.file_obj = uploaded_file
        rules = rules_df.sort_values('Ordre')

        # On récupère l'index du premier message pour ne pas le répéter s'il a déjà servi d'accueil
        first_msg_idx = rules[rules['Action'] == 'MESSAGE'].index[0] if not rules[rules['Action'] == 'MESSAGE'].empty else -1

        for idx, row in rules.iterrows():
            action = row['Action']
            cible = row['Cible']
            params = str(row['Paramètres'])
            on_fail = row['Si Échec']
            msg_excel = str(row['Message / Question'])

            try:
                # --- ACTION : MESSAGE ---
                if action == "MESSAGE":
                    # On évite de répéter le message de bienvenue s'il est déjà affiché
                    if idx == first_msg_idx and len(st.session_state.chat_history) > 0:
                        continue
                    
                    txt = msg_excel
                    for var, val in self.variables.items():
                        txt = txt.replace(f"{{{var}}}", str(val))
                    self.add_message(txt)

                # --- ACTION : FIND_METADATA ---
                elif action == "FIND_METADATA":
                    wb = openpyxl.load_workbook(self.file_obj, data_only=True)
                    ws = wb.active
                    keyword = self.get_param(params, "keyword")
                    found = False
                    for r in range(1, 25): # On scanne un peu plus large
                        for c in range(1, 12):
                            val = ws.cell(r, c).value
                            if val and keyword in str(val):
                                self.variables[cible] = ws.cell(r, c+1).value
                                found = True; break
                        if found: break
                    if found:
                        self.add_message(f"ℹ️ Métadonnée : **{cible}** détectée ({self.variables[cible]})")
                    else: raise Exception(f"Clé '{keyword}' introuvable")

                # --- ACTION : LOAD_DATA_TABLE (Version avec Gestion Optionnelle Robuste) ---
                # --- ACTION : LOAD_DATA_TABLE ---
                elif action == "LOAD_DATA_TABLE":
                    if self.df is not None:
                        # On ne dit rien pour ne pas polluer le chat, on passe juste
                        continue
                    # PROTECTION CRITIQUE : Si on a déjà des données corrigées en mémoire, 
                    # on saute la lecture du fichier Excel pour ne pas écraser les corrections.
                    if self.df is not None:
                        self.add_message(f"🔄 Utilisation des données corrigées en mémoire ({len(self.df)} lignes).")
                        continue 

                    start_key = self.get_param(params, "Start_At")
                    m_param = self.get_param(params, "Mandatory")
                    o_param = self.get_param(params, "Optional")
                    
                    mandatory_raw = [c.strip() for c in m_param.split(';')] if m_param else []
                    optional_raw = [c.strip() for c in o_param.split(';')] if o_param else []
                    
                    temp = pd.read_excel(self.file_obj, header=None, nrows=50)
                    mask = temp.apply(lambda r: r.astype(str).str.contains(start_key, case=False, na=False).any(), axis=1)
                    
                    if not mask.any():
                        raise Exception(f"Clé de départ '{start_key}' introuvable.")
                    
                    header_idx = temp[mask].index[0]
                    actual_columns = pd.read_excel(self.file_obj, skiprows=header_idx, nrows=0).columns.tolist()
                    
                    columns_to_keep = []
                    for m in mandatory_raw:
                        if not m: continue
                        found = self.match_col(m, actual_columns)
                        if found and found not in columns_to_keep: columns_to_keep.append(found)
                        elif not found: raise Exception(f"❌ Colonne obligatoire '{m}' introuvable.")
                            
                    for o in optional_raw:
                        if not o: continue
                        found = self.match_col(o, actual_columns)
                        if found and found not in columns_to_keep: columns_to_keep.append(found)

                    self.file_obj.seek(0)
                    self.df = pd.read_excel(self.file_obj, skiprows=header_idx, usecols=columns_to_keep)
                    
                    # On initialise la mémoire session dès le premier chargement
                    st.session_state.df_principal = self.df
                    
                    self.add_message(f"📊 Tableau chargé : {len(self.df)} lignes.")
                # --- ACTION : RENAME ---
                # --- ACTION : RENAME (Version Sécurisée) ---
                elif action == "RENAME":
                    new_name = self.get_param(params, "To")
                    
                    # SI LA COLONNE A DÉJÀ LE BON NOM (déjà fait au tour précédent)
                    if new_name in self.df.columns:
                        continue # On passe à la suite sans erreur
                    
                    real = self.match_col(cible, self.df.columns)
                    if real:
                        self.df.rename(columns={real: new_name}, inplace=True)
                        self.add_message(f"🏷️ Colonne renommée : `{real}` ➔ `{new_name}`")
                    else:
                        # On ne stoppe que si on n'a ni l'ancien, ni le nouveau nom
                        if on_fail == "STOP":
                            raise Exception(f"Colonne '{cible}' (ou '{new_name}') introuvable.")

                # --- ACTION : INJECT ---
                elif action == "INJECT":
                    var_key = self.get_param(params, "Value")
                    val = self.variables.get(var_key, "N/A")
                    self.df[cible] = val
                    self.add_message(f"💉 Injection : Colonne `{cible}` créée avec la valeur `{val}`")

                # --- ACTION : FILTER ---
                elif action == "FILTER":
                    cond_raw = self.get_param(params, "Condition")
                    
                    # 1. Nettoyage et détection de l'opérateur
                    cond_raw = cond_raw.strip().strip('"').strip("'")
                    found_op = None
                    for op in ['!=', '==', '>=', '<=', '>', '<']:
                        if op in cond_raw:
                            found_op = op
                            break
                    
                    if not found_op:
                        raise Exception(f"Opérateur manquant dans : {cond_raw}")

                    # 2. Séparation Gauche / Droite
                    parts = cond_raw.split(found_op)
                    raw_left = parts[0].strip().replace('`', '')
                    raw_right = parts[1].strip().replace('`', '')

                    # 3. Identification des colonnes réelles dans le DF
                    col_left = self.match_col(raw_left, self.df.columns)
                    col_right = self.match_col(raw_right, self.df.columns)

                    # --- DIAGNOSTIC SI ERREUR ---
                    if not col_left or not col_right:
                        cols_dispo = ", ".join(list(self.df.columns))
                        raise Exception(f"Colonnes introuvables.\nCherché : '{raw_left}' et '{raw_right}'\nDisponible dans le fichier : [{cols_dispo}]")

                    # 4. Filtrage Natif
                    nb_avant = len(self.df)
                    if found_op == '!=':
                        self.df = self.df[self.df[col_left].astype(str).str.strip() != self.df[col_right].astype(str).str.strip()]
                    elif found_op == '==':
                        self.df = self.df[self.df[col_left].astype(str).str.strip() == self.df[col_right].astype(str).str.strip()]
                    
                    # (Note: J'ai ajouté .astype(str).str.strip() pour éviter les erreurs de types ou d'espaces invisibles dans les cellules Excel)
                    
                    self.add_message(f"✂️ Filtrage : {nb_avant} ➔ {len(self.df)} lignes.")

                # --- ACTION : CONTROLE_CORRECT (Version Améliorée avec Correction Interactive) ---
                elif action == "CONTROLE_CORRECT":
                    fmt = self.get_param(params, "Format")
                    real_col = self.match_col(cible, self.df.columns)
                    n = int(fmt.split(':')[1]) if "chars:" in fmt else 3
                    
                    mask_invalid = self.df[real_col].astype(str).str.strip().str.len() != n
                    
                    if mask_invalid.any():
                        st.warning(f"⚠️ `{cible}` : Plusieurs formats sont incorrects.")
                        
                        # On crée un petit DataFrame avec uniquement les valeurs uniques en erreur
                        invalid_df = pd.DataFrame({
                            "Valeur Actuelle": self.df.loc[mask_invalid, real_col].unique()
                        })
                        invalid_df["Nouvelle Valeur (3 chars)"] = "" # Colonne vide à remplir

                        with st.form("bulk_fix"):
                            st.write("Corrigez les correspondances ci-dessous :")
                            # L'utilisateur remplit le tableau
                            edited_df = st.data_editor(invalid_df, use_container_width=True, hide_index=True)
                            submit = st.form_submit_button("Appliquer toutes les corrections")

                            if submit:
                                # On applique chaque correction saisie dans le tableau
                                for _, row_fix in edited_df.iterrows():
                                    old = row_fix["Valeur Actuelle"]
                                    new = row_fix["Nouvelle Valeur (3 chars)"].strip()
                                    
                                    if new.lower() == 'delete':
                                        self.df = self.df[self.df[real_col] != old]
                                    elif len(new) == n:
                                        self.df.loc[self.df[real_col] == old, real_col] = new
                                
                                st.session_state.df_principal = self.df
                                st.rerun()
                        st.stop()

                # --- ACTION : CONTROLE ---
                elif action == "CONTROLE":
                    fmt = self.get_param(params, "Format")
                    
                    # --- ZONE DEBUG AMÉLIORÉE ---
                    with st.expander(f"🛠️ Debug Interne : {cible}", expanded=True):
                        st.write(f"1. Recherche de l'étiquette : `{cible}`")
                        
                        # A. Tentative dans le SAC
                        val_raw = self.variables.get(cible, None)
                        source = "SAC"
                        
                        # B. Si pas dans le sac, tentative dans le TABLEAU (Base)
                        if val_raw is None:
                            if self.df is not None:
                                # On cherche une colonne qui correspond (via match_col pour être sûr)
                                real_col = self.match_col(cible, self.df.columns)
                                if real_col:
                                    val_raw = self.df[real_col].iloc[0] # On prend la 1ère ligne pour le test
                                    source = f"TABLEAU (colonne: {real_col})"
                            
                        if val_raw is None:
                            val_raw = "⚠️ NON TROUVÉ (Ni sac, ni colonne)"
                            source = "AUCUNE"

                        st.write(f"2. Source de la donnée : `{source}`")
                        st.write(f"3. Valeur brute récupérée : `{val_raw}`")
                        
                        val_to_check = str(val_raw).strip()
                        st.write(f"4. Valeur nettoyée pour contrôle : `{val_to_check}`")
                    # ------------------
                    
                    # (Le reste de ton code ne change pas)
                    is_valid = False
                    reason = ""

                    # 1. LOGIQUE DE VÉRIFICATION
                    if "exact_digits:" in fmt:
                        n = int(fmt.split(':')[1])
                        is_valid = val_to_check.isdigit() and len(val_to_check) == n
                        reason = f"{n} digit"
                    elif "alphanum_fixed:" in fmt:
                        n = int(fmt.split(':')[1])
                        is_valid = bool(re.match(rf"^[A-Za-z0-9]{{{n}}}$", val_to_check))
                        reason = f"{n} alphanumerical digits"
                    elif fmt == "alphanum_code":
                        is_valid = bool(re.match(r"^[A-Za-z0-9\s]+$", val_to_check))
                        reason = "alphanumerical code"
                    elif "max_chars:" in fmt:
                        # 1. On nettoie la valeur à tester de façon agressive
                        val_clean = " ".join(val_to_check.split())
                        current_len = len(val_clean)                       
#                        st.write("current_len: ", current_len)

                        try:
                            # 2. On extrait le nombre N du paramètre (ex: "max_chars:28" -> 28)
                            # On ajoute un strip() sur le split pour éviter les erreurs d'espaces
                            n_str = fmt.split(':')[1].strip()
                            n = int(n_str)
                            
                            is_valid = current_len <= n
                            reason = f"max {n} chars (vu: {current_len})"
                        except Exception as e:
                            is_valid = False
                            reason = f"Erreur paramètre : {fmt}"
                            st.error(f"⚠️ Erreur sur le paramètre Format: {fmt}. Détail: {e}")
                    elif "chars:" in fmt:
                        n = int(fmt.split(':')[1])
                        is_valid = len(val_to_check) == n
                        reason = f"{n} characters"
                    # Format : Maximum de caractères (ex: Location)
                    # Format : Maximum de caractères (ex: Location)
                    elif fmt == "numeric_price":
                        val_clean = val_to_check.replace(',', '.')
                        
                        # Accepte : 228 (entier), 228.9 (1 déc), 228.95 (2 déc), 228.951 (3 déc)
                        pattern = r"^\d+(\.\d{1,3})?$"
                        
                        is_valid = bool(re.match(pattern, val_clean))
                        reason = "numeric value (integer or 1-3 decimals)"

                    # 2. AFFICHAGE VISUEL
                    icon = "✅" if is_valid else "❌"
                    color = "green" if is_valid else "red"
                    msg = f"🔍 **{cible}** : {val_to_check} {icon} :{color}[{reason}]"
                    self.add_message(msg)
                    

                    if not is_valid and on_fail == "STOP":
                        raise Exception(f"Validation échouée : {cible} ({reason})")

            except Exception as e:
                self.add_message(f"⚠️ Erreur ({action}): {e}")
                if on_fail == "STOP": return None
        
        return self.df

# --- 4. INTERFACE STREAMLIT ---
st.title("🚀 Harness - Smart Data Engine")

# Chargement du fichier de règles
rules_path = os.path.join(os.path.dirname(__file__), "Rules.xlsx")
if not os.path.exists(rules_path):
    st.error("Fichier Rules.xlsx manquant.")
    st.stop()

rules_df = pd.read_excel(rules_path).sort_values('Ordre')

# --- INITIALISATION DU MESSAGE D'ACCUEIL DEPUIS L'EXCEL ---
if not st.session_state.welcome_done:
    # On prend la première ligne 'MESSAGE' du fichier Excel
    welcome_row = rules_df[rules_df['Action'] == 'MESSAGE'].head(1)
    if not welcome_row.empty:
        st.session_state.chat_history.append(str(welcome_row['Message / Question'].values[0]))
    st.session_state.welcome_done = True

# Zone d'affichage du Chat
st.write("---")
st.subheader("💬 Journal de traitement")
chat_container = st.container()

with chat_container:
    for msg in st.session_state.chat_history:
        with st.chat_message("assistant"):
            st.write(msg)

# Upload et Bouton
data_file = st.file_uploader("📂 Déposez votre fichier Excel ici", type=["xlsx", "xls"])

# --- LOGIQUE DE DÉMARRAGE ET CONTINUITÉ ---
if data_file:
    # 1. On prépare l'analyse au clic sur le bouton
    if st.button("🚀 Démarrer l'analyse"):
        # On réinitialise tout pour un nouveau fichier
        st.session_state.df_principal = None
        st.session_state.chat_history = [st.session_state.chat_history[0]] if st.session_state.chat_history else []
        st.session_state.run_active = True # Un flag pour dire "L'analyse est lancée"

    # 2. Si l'analyse est active (soit via bouton, soit après une correction)
    if st.session_state.get("run_active", False):
        engine = StreamlitDataEngine()
        
        # Si on a déjà des données en mémoire (suite à une correction), on les donne à l'engine
        if st.session_state.df_principal is not None:
            engine.df = st.session_state.df_principal
        
        with chat_container:
            final_df = engine.run_pipeline(rules_df, data_file)

        # 3. Affichage des résultats finaux (seulement quand tout est fini)
        if final_df is not None:
            st.session_state.run_active = False # On a fini, on peut libérer le flag
            st.success("Analyse terminée !")
            st.dataframe(final_df, use_container_width=True)
            
            # Export
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False)
            st.download_button("📥 Télécharger le résultat", output.getvalue(), "Resultat.xlsx")

# ANCRE DE SCROLL FINALE
st.markdown('<div id="end-of-chat"></div>', unsafe_allow_html=True)